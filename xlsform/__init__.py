import click

from openpyxl import Workbook, load_workbook

from .helpers import (
    SurveyHelper,
    ChoicesHelper,
    SettingsHelper
)

def init_workbook(sheet_names):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_names[0]
    for sheet_name in sheet_names[1:]:
        workbook.create_sheet(sheet_name)
    return workbook

@click.command()
@click.option('-m', '--method',  type=click.Choice(['create', 'digest']), required=True)
@click.option('-w', '--workbook_file', required=True, help='path to workbook')
@click.option('-f', '--folder', required=True, help='path to json folder')
def xlsform(method, workbook_file, folder):
    helpers = [
        SurveyHelper(),
        ChoicesHelper(),
        SettingsHelper()
    ]
    if method == 'digest':
        workbook = load_workbook(filename=workbook_file)
        for helper in helpers:
            helper.read_sheet(workbook)
            helper.write_json(folder)
    elif method == 'create':
        workbook = init_workbook([h.SHEET_NAME for h in helpers])
        for helper in helpers:
            helper.read_json(folder)
            helper.write_sheet(workbook)
        workbook.save(filename=workbook_file)
