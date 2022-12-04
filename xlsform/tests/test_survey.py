import unittest

from openpyxl import Workbook

from ..helpers import (
    write_survey,
    read_survey,
    get_columns
)

class TestWriteSurvey(unittest.TestCase):
    def test_basic(self):
        survey = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'select_multiple degree',
                'label': 'How crazy are your plans?',
                'name': 'insanity',
                'hint': 'Honesty is essential'
            }
        ]
        workbook = Workbook()
        sheet = workbook.active
        write_survey(sheet, survey)
        columns = get_columns(sheet)
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            for j, cell in enumerate(row):
                assert cell.value == survey[i].get(columns[j], None)

    def test_single_recursion(self):
        survey = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group',
                'survey': [
                    {
                        'type': 'select_multiple degree',
                        'label': 'How crazy are your plans?',
                        'name': 'insanity',
                        'hint': 'Honesty is essential'
                    }
                ]
            }
        ]
        expected_rows = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group'
            },
            {
                'type': 'select_multiple degree',
                'label': 'How crazy are your plans?',
                'name': 'insanity',
                'hint': 'Honesty is essential'
            },
            {
                'type': 'end group'
            }
        ]
        workbook = Workbook()
        sheet = workbook.active
        write_survey(sheet, survey)
        columns = get_columns(sheet)
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            for j, cell in enumerate(row):
                assert cell.value == expected_rows[i].get(columns[j], None)

    def test_single_recursion_with_underscore(self):
        survey = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin_group',
                'label': 'A Group!',
                'name': 'group',
                'survey': [
                    {
                        'type': 'select_multiple degree',
                        'label': 'How crazy are your plans?',
                        'name': 'insanity',
                        'hint': 'Honesty is essential'
                    }
                ]
            }
        ]
        expected_rows = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin_group',
                'label': 'A Group!',
                'name': 'group'
            },
            {
                'type': 'select_multiple degree',
                'label': 'How crazy are your plans?',
                'name': 'insanity',
                'hint': 'Honesty is essential'
            },
            {
                'type': 'end_group'
            }
        ]
        workbook = Workbook()
        sheet = workbook.active
        write_survey(sheet, survey)
        columns = get_columns(sheet)
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            for j, cell in enumerate(row):
                assert cell.value == expected_rows[i].get(columns[j], None)

    def test_multiple_recursion(self):
        survey = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group',
                'survey': [
                    {
                        'type': 'select_one yes_no',
                        'label': 'Are you a squirrel?',
                        'name': 'is_squirrel'
                    },
                    {
                        'type': 'begin repeat',
                        'survey': [
                            {
                                'type': 'select_multiple degree',
                                'label': 'How crazy are your plans?',
                                'name': 'insanity',
                                'hint': 'Honesty is essential'
                            }
                        ]
                    }
                    
                ]
            }
        ]
        expected_rows = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group'
            },
            {
                'type': 'select_one yes_no',
                'label': 'Are you a squirrel?',
                'name': 'is_squirrel'
            },
            {
                'type': 'begin repeat'
            },
            {
                'type': 'select_multiple degree',
                'label': 'How crazy are your plans?',
                'name': 'insanity',
                'hint': 'Honesty is essential'
            },
            {
                'type': 'end repeat'
            },
            {
                'type': 'end group'
            }
        ]
        workbook = Workbook()
        sheet = workbook.active
        write_survey(sheet, survey)
        columns = get_columns(sheet)
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            for j, cell in enumerate(row):
                assert cell.value == expected_rows[i].get(columns[j], None)

    def test_subsequent_recursion(self):
        survey = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group',
                'survey': [
                    {
                        'type': 'select_one yes_no',
                        'label': 'Are you a squirrel?',
                        'name': 'is_squirrel'
                    } 
                ]
            },
            {
                'type': 'begin repeat',
                'survey': [
                    {
                        'type': 'select_multiple degree',
                        'label': 'How crazy are your plans?',
                        'name': 'insanity',
                        'hint': 'Honesty is essential'
                    }
                ]
            }
        ]
        expected_rows = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group'
            },
            {
                'type': 'select_one yes_no',
                'label': 'Are you a squirrel?',
                'name': 'is_squirrel'
            },
            {
                'type': 'end group'
            },
            {
                'type': 'begin repeat'
            },
            {
                'type': 'select_multiple degree',
                'label': 'How crazy are your plans?',
                'name': 'insanity',
                'hint': 'Honesty is essential'
            },
            {
                'type': 'end repeat'
            }
        ]
        workbook = Workbook()
        sheet = workbook.active
        write_survey(sheet, survey)
        columns = get_columns(sheet)
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            for j, cell in enumerate(row):
                assert cell.value == expected_rows[i].get(columns[j], None)

def recursive_equality(survey, expected_survey):
    assert survey == expected_survey
    for element, expected_element in zip(survey, expected_survey):
        if 'survey' in element or 'survey' in expected_element:
            recursive_equality(
                element['survey'],
                expected_element['survey']
            )
class TestReadSurvey(unittest.TestCase):
    def test_basic(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup the columns
        sheet.cell(row=1, column=1, value='type')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=3, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='select_one yes_no')
        sheet.cell(row=2, column=2, value='is_lizard')
        sheet.cell(row=2, column=3, value='Are you a lizard?')
        # second data row
        sheet.cell(row=3, column=1, value='select_one degree')
        sheet.cell(row=3, column=2, value='insanity')
        sheet.cell(row=3, column=3, value='How crazy are your plans?')

        survey = read_survey(sheet)
        expected_survey = [
            {
                'type': 'select_one yes_no',
                'name': 'is_lizard',
                'label': 'Are you a lizard?'
            },
            {
                'type': 'select_one degree',
                'name': 'insanity',
                'label': 'How crazy are your plans?'
            }
        ]
        recursive_equality(survey, expected_survey)

    def test_single_recursion(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup the columns
        sheet.cell(row=1, column=1, value='type')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=3, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='select_one yes_no')
        sheet.cell(row=2, column=2, value='is_lizard')
        sheet.cell(row=2, column=3, value='Are you a lizard?')
        # second data row
        sheet.cell(row=3, column=1, value='begin group')
        sheet.cell(row=3, column=2, value='group')
        sheet.cell(row=3, column=3, value='A Group!')
        # third data row
        sheet.cell(row=4, column=1, value='select_one degree')
        sheet.cell(row=4, column=2, value='insanity')
        sheet.cell(row=4, column=3, value='How crazy are your plans?')
        # fourth data row
        sheet.cell(row=5, column=1, value='end group')

        survey = read_survey(sheet)
        expected_survey = [
            {
                'type': 'select_one yes_no',
                'name': 'is_lizard',
                'label': 'Are you a lizard?'
            },
            {
                'type': 'begin group',
                'name': 'group',
                'label': 'A Group!',
                'survey': [
                    {
                        'type': 'select_one degree',
                        'name': 'insanity',
                        'label': 'How crazy are your plans?'
                    }
                ]
            }
        ]
        recursive_equality(survey, expected_survey)

    def test_single_recursion_with_underscore(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup the columns
        sheet.cell(row=1, column=1, value='type')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=3, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='select_one yes_no')
        sheet.cell(row=2, column=2, value='is_lizard')
        sheet.cell(row=2, column=3, value='Are you a lizard?')
        # second data row
        sheet.cell(row=3, column=1, value='begin_group')
        sheet.cell(row=3, column=2, value='group')
        sheet.cell(row=3, column=3, value='A Group!')
        # third data row
        sheet.cell(row=4, column=1, value='select_one degree')
        sheet.cell(row=4, column=2, value='insanity')
        sheet.cell(row=4, column=3, value='How crazy are your plans?')
        # fourth data row
        sheet.cell(row=5, column=1, value='end_group')

        survey = read_survey(sheet)
        expected_survey = [
            {
                'type': 'select_one yes_no',
                'name': 'is_lizard',
                'label': 'Are you a lizard?'
            },
            {
                'type': 'begin_group',
                'name': 'group',
                'label': 'A Group!',
                'survey': [
                    {
                        'type': 'select_one degree',
                        'name': 'insanity',
                        'label': 'How crazy are your plans?'
                    }
                ]
            }
        ]
        recursive_equality(survey, expected_survey)

    def test_multiple_recursion(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup the columns
        sheet.cell(row=1, column=1, value='type')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=3, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='select_one yes_no')
        sheet.cell(row=2, column=2, value='is_lizard')
        sheet.cell(row=2, column=3, value='Are you a lizard?')
        # second data row
        sheet.cell(row=3, column=1, value='begin group')
        sheet.cell(row=3, column=2, value='group')
        sheet.cell(row=3, column=3, value='A Group!')
        # third data row
        sheet.cell(row=4, column=1, value='select_one degree')
        sheet.cell(row=4, column=2, value='insanity')
        sheet.cell(row=4, column=3, value='How crazy are your plans?')
        # fourth data row
        sheet.cell(row=5, column=1, value='begin repeat')
        # fifth data row
        sheet.cell(row=6, column=1, value='select_one yes_no')
        sheet.cell(row=6, column=2, value='is_squirrel')
        sheet.cell(row=6, column=3, value='Are you a squirrel?')
        # sixth data row
        sheet.cell(row=7, column=1, value='end repeat')       
        # seventh data row
        sheet.cell(row=8, column=1, value='end group')

        survey = read_survey(sheet)
        expected_survey = [
            {
                'type': 'select_one yes_no',
                'name': 'is_lizard',
                'label': 'Are you a lizard?'
            },
            {
                'type': 'begin group',
                'name': 'group',
                'label': 'A Group!',
                'survey': [
                    {
                        'type': 'select_one degree',
                        'name': 'insanity',
                        'label': 'How crazy are your plans?'
                    },
                    {
                        'type': 'begin repeat',
                        'survey': [
                            {
                                'type': 'select_one yes_no',
                                'name': 'is_squirrel',
                                'label': 'Are you a squirrel?'
                            }
                        ]
                    }
                ]
            }
        ]
        recursive_equality(survey, expected_survey)

    def test_subsequent_recursion(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup the columns
        sheet.cell(row=1, column=1, value='type')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=3, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='select_one yes_no')
        sheet.cell(row=2, column=2, value='is_lizard')
        sheet.cell(row=2, column=3, value='Are you a lizard?')
        # second data row
        sheet.cell(row=3, column=1, value='begin group')
        sheet.cell(row=3, column=2, value='group')
        sheet.cell(row=3, column=3, value='A Group!')
        # third data row
        sheet.cell(row=4, column=1, value='select_one degree')
        sheet.cell(row=4, column=2, value='insanity')
        sheet.cell(row=4, column=3, value='How crazy are your plans?')
        # fourth data row
        sheet.cell(row=5, column=1, value='end group')
        # fifth data row
        sheet.cell(row=6, column=1, value='begin repeat')
        # sixth data row
        sheet.cell(row=7, column=1, value='select_one yes_no')
        sheet.cell(row=7, column=2, value='is_squirrel')
        sheet.cell(row=7, column=3, value='Are you a squirrel?')
        # seventh data row
        sheet.cell(row=8, column=1, value='end repeat')

        survey = read_survey(sheet)
        expected_survey = [
            {
                'type': 'select_one yes_no',
                'name': 'is_lizard',
                'label': 'Are you a lizard?'
            },
            {
                'type': 'begin group',
                'name': 'group',
                'label': 'A Group!',
                'survey': [
                    {
                        'type': 'select_one degree',
                        'name': 'insanity',
                        'label': 'How crazy are your plans?'
                    }
                ]
            },
            {
                'type': 'begin repeat',
                'survey': [
                    {
                        'type': 'select_one yes_no',
                        'name': 'is_squirrel',
                        'label': 'Are you a squirrel?'
                    }
                ]
            }
        ]
        recursive_equality(survey, expected_survey)

    def test_empty_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup the columns
        sheet.cell(row=1, column=1, value='type')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=3, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='select_one yes_no')
        sheet.cell(row=2, column=2, value='is_lizard')
        sheet.cell(row=2, column=3, value='Are you a lizard?')
        # second data row
        sheet.cell(row=4, column=1, value='select_one degree')
        sheet.cell(row=4, column=2, value='insanity')
        sheet.cell(row=4, column=3, value='How crazy are your plans?')

        survey = read_survey(sheet)
        expected_survey = [
            {
                'type': 'select_one yes_no',
                'name': 'is_lizard',
                'label': 'Are you a lizard?'
            },
            {
                'type': 'select_one degree',
                'name': 'insanity',
                'label': 'How crazy are your plans?'
            }
        ]
        recursive_equality(survey, expected_survey)

    def test_empty_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup the columns
        sheet.cell(row=1, column=1, value='type')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=4, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='select_one yes_no')
        sheet.cell(row=2, column=2, value='is_lizard')
        sheet.cell(row=2, column=4, value='Are you a lizard?')
        # second data row
        sheet.cell(row=3, column=1, value='select_one degree')
        sheet.cell(row=3, column=2, value='insanity')
        sheet.cell(row=3, column=4, value='How crazy are your plans?')

        survey = read_survey(sheet)
        expected_survey = [
            {
                'type': 'select_one yes_no',
                'name': 'is_lizard',
                'label': 'Are you a lizard?'
            },
            {
                'type': 'select_one degree',
                'name': 'insanity',
                'label': 'How crazy are your plans?'
            }
        ]
        recursive_equality(survey, expected_survey)
