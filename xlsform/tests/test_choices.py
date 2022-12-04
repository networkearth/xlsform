import unittest

from openpyxl import Workbook

from ..helpers import (
    write_choices,
    read_choices,
    get_columns,
    get_extra_keys
)

class TestGetExtraKeys(unittest.TestCase):
    def test_basic(self):
        choices = {
            'yes_no': {
                'yes': {'label': 'Yes', 'media::image::English (en)': ''},
                'no': {'label': 'No', 'media::image::English (en)': ''}
            },
            'degree': {
                'low': {'label': 'Low', 'media::image::English (en)': 'incidental.png'}
            }
        }
        extra_keys = get_extra_keys(choices)
        assert set(extra_keys.keys()) == set(['label', 'media::image::English (en)'])
        assert set(extra_keys.values()) == set([0, 1])

class TestWriteChoices(unittest.TestCase):
    def test_basic(self):
        choices = {
            'yes_no': {
                'yes': {'label': 'Yes'},
                'no': {'label': 'No'}
            },
            'degree': {
                'low': {'label': 'Low'},
                'medium': {'label': 'Medium'},
                'high': {'label': 'High'}
            }
        }
        workbook = Workbook()
        sheet = workbook.active
        write_choices(sheet, choices)
        columns = {
            value: key for key, value in get_columns(sheet).items()
        }
        for row in sheet.iter_rows(min_row=2):
            key = row[columns['list_name']].value
            name = row[columns['name']].value
            assert row[columns['label']].value == choices[key][name]['label']

    def test_basic_extra_data(self):
        choices = {
            'yes_no': {
                'yes': {'label': 'Yes', 'extra': 'stuff'},
                'no': {'label': 'No', 'extra': 'things'}
            },
            'degree': {
                'low': {'label': 'Low', 'extra': ''},
                'medium': {'label': 'Medium', 'extra': ''},
                'high': {'label': 'High', 'extra': ''}
            }
        }
        workbook = Workbook()
        sheet = workbook.active
        write_choices(sheet, choices)
        columns = {
            value: key for key, value in get_columns(sheet).items()
        }
        for row in sheet.iter_rows(min_row=2):
            key = row[columns['list_name']].value
            name = row[columns['name']].value
            assert row[columns['label']].value == choices[key][name]['label']
            assert row[columns['extra']].value == choices[key][name]['extra']

class TestReadChoices(unittest.TestCase):
    def test_basic(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup columns
        sheet.cell(row=1, column=1, value='list_name')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=3, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='yes_no')
        sheet.cell(row=2, column=2, value='yes')
        sheet.cell(row=2, column=3, value='Yes')
        # second data row
        sheet.cell(row=3, column=1, value='yes_no')
        sheet.cell(row=3, column=2, value='no')
        sheet.cell(row=3, column=3, value='No')
        # third data row
        sheet.cell(row=4, column=1, value='degree')
        sheet.cell(row=4, column=2, value='low')
        sheet.cell(row=4, column=3, value='Low')

        choices = read_choices(sheet)
        expected_choices = {
            'yes_no': {
                'yes': {'label': 'Yes'},
                'no': {'label': 'No'}
            },
            'degree': {
                'low': {'label': 'Low'}
            }
        }
        assert choices == expected_choices

    def test_empty_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup columns
        sheet.cell(row=1, column=1, value='list_name')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=3, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='yes_no')
        sheet.cell(row=2, column=2, value='yes')
        sheet.cell(row=2, column=3, value='Yes')
        # second data row
        sheet.cell(row=3, column=1, value='yes_no')
        sheet.cell(row=3, column=2, value='no')
        sheet.cell(row=3, column=3, value='No')
        # third data row
        sheet.cell(row=5, column=1, value='degree')
        sheet.cell(row=5, column=2, value='low')
        sheet.cell(row=5, column=3, value='Low')

        choices = read_choices(sheet)
        expected_choices = {
            'yes_no': {
                'yes': {'label': 'Yes'},
                'no': {'label': 'No'}
            },
            'degree': {
                'low': {'label': 'Low'}
            }
        }
        assert choices == expected_choices

    def test_empty_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup columns
        sheet.cell(row=1, column=1, value='list_name')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=4, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='yes_no')
        sheet.cell(row=2, column=2, value='yes')
        sheet.cell(row=2, column=4, value='Yes')
        # second data row
        sheet.cell(row=3, column=1, value='yes_no')
        sheet.cell(row=3, column=2, value='no')
        sheet.cell(row=3, column=4, value='No')
        # third data row
        sheet.cell(row=4, column=1, value='degree')
        sheet.cell(row=4, column=2, value='low')
        sheet.cell(row=4, column=4, value='Low')

        choices = read_choices(sheet)
        expected_choices = {
            'yes_no': {
                'yes': {'label': 'Yes'},
                'no': {'label': 'No'}
            },
            'degree': {
                'low': {'label': 'Low'}
            }
        }
        assert choices == expected_choices

    def test_extra_info(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup columns
        sheet.cell(row=1, column=1, value='list_name')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=3, value='label')
        sheet.cell(row=1, column=4, value='media::image::English (en)')
        # first data row
        sheet.cell(row=2, column=1, value='yes_no')
        sheet.cell(row=2, column=2, value='yes')
        sheet.cell(row=2, column=3, value='Yes')
        # second data row
        sheet.cell(row=3, column=1, value='yes_no')
        sheet.cell(row=3, column=2, value='no')
        sheet.cell(row=3, column=3, value='No')
        # third data row
        sheet.cell(row=4, column=1, value='degree')
        sheet.cell(row=4, column=2, value='low')
        sheet.cell(row=4, column=3, value='Low')
        sheet.cell(row=4, column=4, value='incidental.png')

        choices = read_choices(sheet)
        expected_choices = {
            'yes_no': {
                'yes': {'label': 'Yes', 'media::image::English (en)': ''},
                'no': {'label': 'No', 'media::image::English (en)': ''}
            },
            'degree': {
                'low': {'label': 'Low', 'media::image::English (en)': 'incidental.png'}
            }
        }
        assert choices == expected_choices