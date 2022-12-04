from cmath import exp
import unittest

from openpyxl import Workbook

from ..helpers import (
    write_settings,
    read_settings
)

class TestWriteSettings(unittest.TestCase):
    def test_basic(self):
        settings = {
            'title': 'A Survey',
            'style': 'grid'
        }
        workbook = Workbook()
        sheet = workbook.active
        write_settings(sheet, settings)
        for column in sheet.columns:
            key = column[0].value
            assert settings[key] == column[1].value

class TestReadSettings(unittest.TestCase):
    def test_basic(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup columns
        sheet.cell(row=1, column=1, value='title')
        sheet.cell(row=1, column=2, value='style')
        # data
        sheet.cell(row=2, column=1, value='A Survey')
        sheet.cell(row=2, column=2, value='grid')

        settings = read_settings(sheet)
        expected_settings = {
            'title': 'A Survey',
            'style': 'grid'
        }
        assert settings == expected_settings

    def test_empty_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup columns
        sheet.cell(row=1, column=1, value='title')
        sheet.cell(row=1, column=2, value='style')
        # data
        sheet.cell(row=3, column=1, value='A Survey')
        sheet.cell(row=3, column=2, value='grid')

        settings = read_settings(sheet)
        expected_settings = {
            'title': 'A Survey',
            'style': 'grid'
        }
        assert settings == expected_settings

    def test_empty_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup columns
        sheet.cell(row=1, column=1, value='title')
        sheet.cell(row=1, column=3, value='style')
        # data
        sheet.cell(row=2, column=1, value='A Survey')
        sheet.cell(row=2, column=3, value='grid')

        settings = read_settings(sheet)
        expected_settings = {
            'title': 'A Survey',
            'style': 'grid'
        }
        assert settings == expected_settings

    def test_empty_setting(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup columns
        sheet.cell(row=1, column=1, value='title')
        sheet.cell(row=1, column=3, value='style')
        # data
        sheet.cell(row=2, column=1, value='A Survey')

        settings = read_settings(sheet)
        expected_settings = {
            'title': 'A Survey'
        }
        assert settings == expected_settings
